# dunning/exporters.py
import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import io
from datetime import date
from typing import Optional


# --- PDF Export ---

def render_pdf(ledger: dict) -> bytes:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER
    except ImportError:
        raise RuntimeError("reportlab not installed. Run: pip install reportlab")

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story  = []

    # --- Header ---
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, spaceAfter=4)
    sub_style   = ParagraphStyle("sub",   parent=styles["Normal"],   fontSize=9,  textColor=colors.grey)
    story.append(Paragraph("Account Statement", title_style))
    story.append(Paragraph(f"Customer: {ledger['customer_name']} ({ledger['customer_code']})", styles["Normal"]))
    story.append(Paragraph(f"Statement Date: {ledger['statement_date']}", sub_style))
    if ledger.get("reporting_currency"):
        story.append(Paragraph(f"Reporting Currency: {ledger['reporting_currency']}", sub_style))
    story.append(Spacer(1, 6*mm))

    # --- Invoice Table ---
    show_reporting = bool(ledger.get("reporting_currency"))
    headers = ["Invoice #", "Inv Date", "Due Date", "CCY", "Outstanding", "DPD", "Status"]
    if show_reporting:
        headers += ["Rate", f"Amt ({ledger['reporting_currency']})"]

    rows = [headers]
    for inv in ledger["invoices"]:
        row = [
            inv["invoice_number"],
            inv["invoice_date"],
            inv["due_date"],
            inv["billing_currency"],
            f"{inv['outstanding_amount']:,.2f}",
            str(inv["dpd"]),
            inv["status"].capitalize(),
        ]
        if show_reporting:
            row.append(f"{inv['exchange_rate']:.4f}" if inv.get("exchange_rate") else "N/A")
            row.append(f"{inv['reporting_amount']:,.2f}" if inv.get("reporting_amount") else "N/A")
        rows.append(row)

    col_widths = [35*mm, 22*mm, 22*mm, 14*mm, 28*mm, 12*mm, 18*mm]
    if show_reporting:
        col_widths += [18*mm, 28*mm]

    tbl = Table(rows, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0),  colors.HexColor("#2d6a9f")),
        ("TEXTCOLOR",    (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",     (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),  8),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("ALIGN",        (4, 0), (4, -1),  "RIGHT"),
        ("ALIGN",        (5, 0), (5, -1),  "CENTER"),
        ("ALIGN",        (-1, 0), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f8fc")]),
        ("GRID",         (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6*mm))

    # --- Totals ---
    totals_data = []
    for ccy, amt in ledger["total_by_currency"].items():
        totals_data.append([f"Total ({ccy}):", f"{amt:,.2f}"])
    if show_reporting and ledger.get("total_reporting") is not None:
        totals_data.append([
            f"Total ({ledger['reporting_currency']}):",
            f"{ledger['total_reporting']:,.2f}"
        ])

    if totals_data:
        tot_tbl = Table(totals_data, colWidths=[50*mm, 35*mm])
        tot_tbl.setStyle(TableStyle([
            ("FONTNAME",  (0, 0), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE",  (0, 0), (-1, -1), 9),
            ("ALIGN",     (1, 0), (1, -1),  "RIGHT"),
            ("LINEABOVE", (0, 0), (-1, 0),  0.5, colors.black),
        ]))
        story.append(tot_tbl)

    if ledger.get("rate_missing"):
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(
            "* Some exchange rates not available — reporting amounts may be incomplete.",
            ParagraphStyle("warn", parent=styles["Normal"], fontSize=8, textColor=colors.red)
        ))

    doc.build(story)
    return buf.getvalue()


# --- Excel Export ---

def render_excel(ledger: dict) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AR Statement"

    header_fill  = PatternFill("solid", fgColor="2d6a9f")
    header_font  = Font(bold=True, color="FFFFFF", size=9)
    alt_fill     = PatternFill("solid", fgColor="f5f8fc")
    bold_font    = Font(bold=True, size=9)
    normal_font  = Font(size=9)
    center_align = Alignment(horizontal="center")
    right_align  = Alignment(horizontal="right")
    thin_border  = Border(
        left=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
        top=Side(style="thin", color="cccccc"),
        bottom=Side(style="thin", color="cccccc"),
    )

    # --- Title block ---
    ws["A1"] = "Account Statement"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Customer: {ledger['customer_name']} ({ledger['customer_code']})"
    ws["A2"].font = normal_font
    ws["A3"] = f"Statement Date: {ledger['statement_date']}"
    ws["A3"].font = normal_font
    if ledger.get("reporting_currency"):
        ws["A4"] = f"Reporting Currency: {ledger['reporting_currency']}"
        ws["A4"].font = normal_font

    ws.append([])

    # --- Column headers ---
    show_reporting = bool(ledger.get("reporting_currency"))
    headers = ["Invoice #", "Invoice Date", "Due Date", "Currency", "Outstanding", "DPD", "Status"]
    if show_reporting:
        headers += ["Exchange Rate", f"Amount ({ledger['reporting_currency']})"]

    ws.append(headers)
    hdr_row = ws.max_row
    for col, _ in enumerate(headers, 1):
        cell            = ws.cell(row=hdr_row, column=col)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = center_align
        cell.border     = thin_border

    # --- Data rows ---
    for i, inv in enumerate(ledger["invoices"]):
        row = [
            inv["invoice_number"],
            inv["invoice_date"],
            inv["due_date"],
            inv["billing_currency"],
            inv["outstanding_amount"],
            inv["dpd"],
            inv["status"].capitalize(),
        ]
        if show_reporting:
            row.append(inv.get("exchange_rate") or "N/A")
            row.append(inv.get("reporting_amount") or "N/A")
        ws.append(row)

        fill = alt_fill if i % 2 == 1 else None
        data_row = ws.max_row
        for col in range(1, len(headers) + 1):
            cell           = ws.cell(row=data_row, column=col)
            cell.font      = normal_font
            cell.border    = thin_border
            if fill:
                cell.fill  = fill
            if col in (5, 9):
                cell.alignment = right_align
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
            if col == 6:
                cell.alignment = center_align

    ws.append([])

    # --- Totals ---
    for ccy, amt in ledger["total_by_currency"].items():
        ws.append([f"Total ({ccy})", "", "", "", amt, "", ""])
        tot_row = ws.max_row
        ws.cell(row=tot_row, column=1).font = bold_font
        ws.cell(row=tot_row, column=5).font = bold_font
        ws.cell(row=tot_row, column=5).number_format = "#,##0.00"
        ws.cell(row=tot_row, column=5).alignment = right_align

    if show_reporting and ledger.get("total_reporting") is not None:
        ws.append([f"Total ({ledger['reporting_currency']})", "", "", "", "", "", "",
                   "", ledger["total_reporting"]])
        tot_row = ws.max_row
        ws.cell(row=tot_row, column=1).font = bold_font
        ws.cell(row=tot_row, column=9).font = bold_font
        ws.cell(row=tot_row, column=9).number_format = "#,##0.00"
        ws.cell(row=tot_row, column=9).alignment = right_align

    if ledger.get("rate_missing"):
        ws.append([])
        ws.append(["* Some exchange rates not available — reporting amounts may be incomplete."])
        warn_row = ws.max_row
        ws.cell(row=warn_row, column=1).font = Font(size=8, color="FF0000")

    # --- Column widths ---
    col_widths = [20, 14, 14, 10, 16, 8, 14, 14, 18]
    for i, w in enumerate(col_widths[:len(headers)], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



